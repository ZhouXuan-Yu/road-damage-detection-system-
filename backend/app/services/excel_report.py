"""
Excel 报告生成模块
使用 openpyxl 生成包含多个工作表的病害检测分析报告
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models.detection import DetectionRecord
from app.core.constants import DAMAGE_NAME_CN


def create_styles():
    """创建样式定义"""
    # 表头样式
    header_fill = PatternFill("solid", fgColor="1e40af")
    header_font = Font(name="SimHei", size=11, bold=True, color="FFFFFF")
    
    # 标签样式
    label_font = Font(name="SimHei", size=10, bold=True)
    
    # 值样式
    value_font = Font(name="SimHei", size=10)
    
    # 边框样式
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    # 交替行样式
    alt_fill = PatternFill("solid", fgColor="f9fafb")
    
    return {
        "header_fill": header_fill,
        "header_font": header_font,
        "label_font": label_font,
        "value_font": value_font,
        "thin_border": thin_border,
        "alt_fill": alt_fill,
    }


def apply_header_style(cell, styles: dict):
    """应用表头样式"""
    cell.font = styles["header_font"]
    cell.fill = styles["header_fill"]
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = styles["thin_border"]


def apply_value_style(cell, styles: dict, center: bool = True, alt_bg: bool = False):
    """应用值样式"""
    cell.font = styles["value_font"]
    if alt_bg:
        cell.fill = styles["alt_fill"]
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    cell.border = styles["thin_border"]


def apply_label_style(cell, styles: dict):
    """应用标签样式"""
    cell.font = styles["label_font"]
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = styles["thin_border"]


def generate_excel_report(
    records: list[DetectionRecord],
    summary: dict,
    output_path: Path,
) -> str:
    """
    生成 Excel 报告，包含多个工作表
    
    Args:
        records: 检测记录列表
        summary: 汇总数据
        output_path: 输出文件路径
    
    Returns:
        str: 报告摘要
    """
    wb = Workbook()
    styles = create_styles()
    
    # ==================== Sheet1: 检测概况 ====================
    ws1 = wb.active
    ws1.title = "检测概况"
    
    # 设置列宽
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 35
    
    # 标题
    ws1.merge_cells("A1:B1")
    ws1["A1"] = "道路病害检测分析报告"
    ws1["A1"].font = Font(name="SimHei", size=16, bold=True, color="1e40af")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30
    
    # 报告信息
    report_id = f"REP-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    overview_rows = [
        ("报告编号", report_id),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("报告类型", "Excel 工作簿"),
        ("包含检测记录", f"{len(records)} 条"),
        ("病害目标总数", f"{summary['total_objects']} 个"),
        ("平均置信度", f"{summary['avg_confidence']:.2%}"),
    ]
    
    for i, (label, value) in enumerate(overview_rows, 3):
        ws1.cell(i, 1, label)
        apply_label_style(ws1.cell(i, 1), styles)
        
        ws1.cell(i, 2, value)
        apply_value_style(ws1.cell(i, 2), styles, center=False)
    
    # ==================== Sheet2: 病害分布 ====================
    ws2 = wb.create_sheet("病害分布")
    
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 12
    
    # 标题
    ws2.merge_cells("A1:E1")
    ws2["A1"] = "病害类型分布统计"
    ws2["A1"].font = Font(name="SimHei", size=14, bold=True, color="1e40af")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 25
    
    # 表头
    headers = ["序号", "病害类型", "代码", "数量", "占比"]
    for j, header in enumerate(headers, 1):
        cell = ws2.cell(3, j, header)
        apply_header_style(cell, styles)
    
    # 数据
    total = summary["total_objects"]
    sorted_dist = sorted(
        summary["damage_distribution"].items(),
        key=lambda x: -x[1]
    )
    
    for i, (cls, count) in enumerate(sorted_dist):
        row_idx = i + 4
        pct = count / total * 100 if total > 0 else 0
        
        ws2.cell(row_idx, 1, i + 1)
        apply_value_style(ws2.cell(row_idx, 1), styles, alt_bg=(i % 2 == 1))
        
        ws2.cell(row_idx, 2, DAMAGE_NAME_CN.get(cls, cls))
        apply_value_style(ws2.cell(row_idx, 2), styles, center=False, alt_bg=(i % 2 == 1))
        
        ws2.cell(row_idx, 3, cls)
        apply_value_style(ws2.cell(row_idx, 3), styles, alt_bg=(i % 2 == 1))
        
        ws2.cell(row_idx, 4, count)
        apply_value_style(ws2.cell(row_idx, 4), styles, alt_bg=(i % 2 == 1))
        
        ws2.cell(row_idx, 5, f"{pct:.1f}%")
        apply_value_style(ws2.cell(row_idx, 5), styles, alt_bg=(i % 2 == 1))
    
    # ==================== Sheet3: 严重程度分布 ====================
    if summary.get("severity_distribution"):
        ws3 = wb.create_sheet("严重程度")
        
        ws3.column_dimensions["A"].width = 15
        ws3.column_dimensions["B"].width = 12
        ws3.column_dimensions["C"].width = 25
        
        # 标题
        ws3.merge_cells("A1:C1")
        ws3["A1"] = "病害严重程度分布"
        ws3["A1"].font = Font(name="SimHei", size=14, bold=True, color="1e40af")
        ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[1].height = 25
        
        # 表头
        sev_headers = ["严重程度", "数量", "说明"]
        for j, header in enumerate(sev_headers, 1):
            cell = ws3.cell(3, j, header)
            apply_header_style(cell, styles)
        
        # 数据
        sev_map = {"low": "轻度", "medium": "中度", "high": "高"}
        sev_desc = {
            "low": "面积占比 < 1%",
            "medium": "面积占比 1-5%",
            "high": "面积占比 > 5%"
        }
        sev_colors = {
            "low": "dcfce7",      # 绿色
            "medium": "fef9c3",   # 黄色
            "high": "fee2e2",     # 红色
        }
        
        for i, (sev, count) in enumerate(summary["severity_distribution"].items()):
            row_idx = i + 4
            
            ws3.cell(row_idx, 1, sev_map.get(sev, sev))
            apply_value_style(ws3.cell(row_idx, 1), styles)
            
            ws3.cell(row_idx, 2, count)
            apply_value_style(ws3.cell(row_idx, 2), styles)
            
            ws3.cell(row_idx, 3, sev_desc.get(sev, ""))
            apply_value_style(ws3.cell(row_idx, 3), styles, center=False)
            
            # 设置背景色
            fill = PatternFill("solid", fgColor=sev_colors.get(sev, "ffffff"))
            for j in range(1, 4):
                ws3.cell(row_idx, j).fill = fill
    
    # ==================== Sheet4: 详细记录 ====================
    ws4 = wb.create_sheet("详细记录")
    
    ws4.column_dimensions["A"].width = 8
    ws4.column_dimensions["B"].width = 35
    ws4.column_dimensions["C"].width = 10
    ws4.column_dimensions["D"].width = 12
    ws4.column_dimensions["E"].width = 12
    ws4.column_dimensions["F"].width = 20
    
    # 标题
    ws4.merge_cells("A1:F1")
    ws4["A1"] = "检测记录详情"
    ws4["A1"].font = Font(name="SimHei", size=14, bold=True, color="1e40af")
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 25
    
    # 表头
    detail_headers = ["序号", "文件名", "类型", "病害数", "置信度", "检测时间"]
    for j, header in enumerate(detail_headers, 1):
        cell = ws4.cell(3, j, header)
        apply_header_style(cell, styles)
    
    # 数据
    for i, r in enumerate(records):
        row_idx = i + 4
        
        ws4.cell(row_idx, 1, i + 1)
        apply_value_style(ws4.cell(row_idx, 1), styles, alt_bg=(i % 2 == 1))
        
        ws4.cell(row_idx, 2, r.filename)
        apply_value_style(ws4.cell(row_idx, 2), styles, center=False, alt_bg=(i % 2 == 1))
        
        ws4.cell(row_idx, 3, "图片" if r.file_type == "image" else "视频")
        apply_value_style(ws4.cell(row_idx, 3), styles, alt_bg=(i % 2 == 1))
        
        ws4.cell(row_idx, 4, r.total_count)
        apply_value_style(ws4.cell(row_idx, 4), styles, alt_bg=(i % 2 == 1))
        
        ws4.cell(row_idx, 5, f"{r.avg_confidence:.2%}")
        apply_value_style(ws4.cell(row_idx, 5), styles, alt_bg=(i % 2 == 1))
        
        ts = r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "-"
        ws4.cell(row_idx, 6, ts)
        apply_value_style(ws4.cell(row_idx, 6), styles, alt_bg=(i % 2 == 1))
    
    # ==================== Sheet5: 原始数据 ====================
    ws5 = wb.create_sheet("原始数据")
    
    ws5.column_dimensions["A"].width = 8
    ws5.column_dimensions["B"].width = 10
    ws5.column_dimensions["C"].width = 30
    ws5.column_dimensions["D"].width = 15
    ws5.column_dimensions["E"].width = 15
    ws5.column_dimensions["F"].width = 12
    ws5.column_dimensions["G"].width = 12
    
    # 标题
    ws5.merge_cells("A1:G1")
    ws5["A1"] = "病害事件原始数据"
    ws5["A1"].font = Font(name="SimHei", size=14, bold=True, color="1e40af")
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 25
    
    # 表头
    raw_headers = ["序号", "记录ID", "文件名", "病害代码", "病害名称", "置信度", "严重程度"]
    for j, header in enumerate(raw_headers, 1):
        cell = ws5.cell(3, j, header)
        apply_header_style(cell, styles)
    
    # 数据
    row_idx = 4
    for i, record in enumerate(records):
        for j, occ in enumerate(record.occurrences or []):
            ws5.cell(row_idx, 1, row_idx - 3)
            apply_value_style(ws5.cell(row_idx, 1), styles, alt_bg=((row_idx - 4) % 2 == 1))
            
            ws5.cell(row_idx, 2, record.id)
            apply_value_style(ws5.cell(row_idx, 2), styles, alt_bg=((row_idx - 4) % 2 == 1))
            
            ws5.cell(row_idx, 3, record.filename)
            apply_value_style(ws5.cell(row_idx, 3), styles, center=False, alt_bg=((row_idx - 4) % 2 == 1))
            
            ws5.cell(row_idx, 4, occ.class_code)
            apply_value_style(ws5.cell(row_idx, 4), styles, alt_bg=((row_idx - 4) % 2 == 1))
            
            ws5.cell(row_idx, 5, DAMAGE_NAME_CN.get(occ.class_code, occ.class_name))
            apply_value_style(ws5.cell(row_idx, 5), styles, center=False, alt_bg=((row_idx - 4) % 2 == 1))
            
            ws5.cell(row_idx, 6, f"{occ.confidence:.4f}")
            apply_value_style(ws5.cell(row_idx, 6), styles, alt_bg=((row_idx - 4) % 2 == 1))
            
            ws5.cell(row_idx, 7, occ.severity or "-")
            apply_value_style(ws5.cell(row_idx, 7), styles, alt_bg=((row_idx - 4) % 2 == 1))
            
            row_idx += 1
    
    # 保存文件
    wb.save(str(output_path))
    
    # 计算总明细数
    total_occurrences = sum(len(r.occurrences or []) for r in records)
    return f"Excel 报告生成成功，共 {len(records)} 条记录，{total_occurrences} 条明细数据"
